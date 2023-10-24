import openpyxl
from collections import defaultdict

# Load the workbook and sheet
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# Initialize dictionaries using defaultdict
products_per_supplier = defaultdict(int)
total_value_per_supplier = defaultdict(float)
products_under_10_inv = {}

# Initialize a list for batch cell updates
cell_updates = []

# Iterate through the rows
for product_row in range(2, product_list.max_row + 1):
    product_num = product_list.cell(product_row, 1).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    supplier_name = product_list.cell(product_row, 4).value
    inventory_price = inventory * price  # Calculate once

    # Update dictionaries using defaultdict
    products_per_supplier[supplier_name] += 1
    total_value_per_supplier[supplier_name] += inventory_price

    # Add to products_under_10_inv if inventory is less than 10
    if inventory < 10:
        products_under_10_inv[product_num] = inventory

    # Add cell update to the list
    cell_updates.append((product_row, 5, inventory_price))

# Batch update the cells
for row, col, value in cell_updates:
    product_list.cell(row, col, value=value)

# Save the workbook
inv_file.save("inventory_with_total_value.xlsx")

# Print the results
print(dict(products_per_supplier))
print(dict(total_value_per_supplier))
print(products_under_10_inv)
